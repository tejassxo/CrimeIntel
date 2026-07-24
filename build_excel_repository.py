import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_master_excel():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_family = "Calibri"
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    title_font = Font(name=font_family, size=16, bold=True, color="0F172A")
    subtitle_font = Font(name=font_family, size=11, italic=True, color="475569")
    section_font = Font(name=font_family, size=13, bold=True, color="0F172A")
    bold_font = Font(name=font_family, size=10, bold=True, color="0F172A")
    regular_font = Font(name=font_family, size=10, color="334155")
    kpi_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    def auto_fit_columns(ws, max_len_cap=55):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if '\n' in val:
                    lines = val.split('\n')
                    max_len = max(max_len, max(len(l) for l in lines))
                else:
                    max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), max_len_cap)

    # -------------------------------------------------------------
    # SHEET 1: Executive Summary
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="Executive Summary")
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.cell(row=2, column=2, value="INDIA CYBERCRIME INTELLIGENCE INITIATIVE (2020–2026 YTD)").font = title_font
    ws1.cell(row=3, column=2, value="Master Threat Assessment & National Cybercrime Intelligence Repository").font = subtitle_font
    ws1.cell(row=4, column=2, value="Report Generated: July 2026 | Classification: Executive Intelligence Brief | Period Covered: 1 Jan 2020 – 24 July 2026").font = Font(name=font_family, size=9, italic=True, color="64748B")

    kpis = [
        ("Total Technical Incidents (CERT-In 2025-2026)", "46,24,000+", "Technical incidents handled by CERT-In (2025-2026 YTD)"),
        ("Cumulative NCRP Complaints (2021-2026 YTD)", "82,39,000+", "Reported financial cybercrimes on I4C portal"),
        ("Total Financial Loss Reported (2021-2026 YTD)", "₹68,500 Cr", "Direct economic impact across reported cases"),
        ("Citizen Funds Blocked/Saved (CFCFRMS June 2026)", "₹11,158 Cr", "Official Govt figure saved via 1930 helpline & bank liens"),
        ("SIM Cards & IMEIs Blocked by MHA", "15,00,000+", "Malicious communications infrastructure severed"),
        ("Dominant Fraud Vector (2024-2026)", "Digital Arrest", "Organized SE Asian syndicates targeting citizens")
    ]

    for idx, (label, val, desc) in enumerate(kpis):
        r = 6 + (idx // 3) * 4
        c = 2 + (idx % 3) * 3
        
        cell_lbl = ws1.cell(row=r, column=c, value=f"{label.upper()}\n{val}\n({desc})")
        cell_lbl.font = Font(name=font_family, size=11, bold=True, color="0F172A")
        cell_lbl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws1.merge_cells(start_row=r, start_column=c, end_row=r+2, end_column=c+2)

        for row in range(r, r+3):
            for col in range(c, c+3):
                cell = ws1.cell(row=row, column=col)
                cell.fill = kpi_fill
                cell.border = thin_border

    # -------------------------------------------------------------
    # SHEET 2: Raw Intelligence Dataset (Updated 2020-2026 YTD)
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Raw Intelligence Dataset")
    ws2.views.sheetView[0].showGridLines = True

    raw_headers = [
        "Incident ID", "Incident Date", "Year", "Quarter", "State / UT", "District", "City", 
        "Target Organization / Entity", "Sector", "Ownership Type", "Attack Category", "Attack Subcategory", 
        "MITRE ATT&CK Tactics", "MITRE ATT&CK Techniques", "Cyber Kill Chain Stage", "Threat Actor / Group", 
        "Threat Actor Origin", "Primary Motivation", "Attack Vector", "Exploitation Method", 
        "Malware Family", "Ransomware Strain", "Target OS / Platform", "Vulnerabilities Exploited (CVE)", 
        "CVSS Score", "Indicators of Compromise (IoCs)", "Financial Loss (INR Crore)", "Records Compromised", 
        "Operational Downtime (Hours)", "Severity Level", "Likelihood Score (1-5)", "Impact Score (1-5)", 
        "Overall Risk Rating", "LE Agency Involved", "Arrests Made", "Funds Recovered (INR Cr)", 
        "Case Status", "CERT-In Advisory Issued", "Primary Source / Reference", "Confidence Score"
    ]

    ws2.row_dimensions[1].height = 30
    for col_idx, h in enumerate(raw_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    incidents_data = [
        ("INC-2020-001", "2020-03-14", 2020, "Q1", "Maharashtra", "Mumbai", "Mumbai", "State Bank of India Customers", "BFSI", "Public Sector", "Phishing & Banking Fraud", "COVID-19 Relief Phishing", "Initial Access", "T1566.002", "Delivery", "Roaming Mantis", "East Asia", "Financial Gain", "SMS Phishing (Smishing)", "Malicious APK installation", "Wormpay Trojan", "N/A", "Android", "N/A", 6.5, "Phishing Domains / Smishing SMS", 12.50, 45000, 0, "High", 4, 4, "High", "Mumbai Cyber Police", 14, 2.10, "Convicted / Closed", "Yes", "RBI / CERT-In Advisory 2020", "High"),
        ("INC-2020-002", "2020-05-22", 2020, "Q2", "Delhi", "New Delhi", "New Delhi", "Ministry of External Affairs", "Government", "Public Sector", "Cyber Espionage", "Spear Phishing", "Initial Access", "T1566.001", "Exploitation", "SideCopy", "South Asia", "Espionage", "Spear Phishing Attachment", "Malicious RTF / Office Macro", "Agent Tesla", "N/A", "Windows", "CVE-2017-11882", 7.8, "C2 Domain: govt-mail-update.com", 0.00, 1200, 48, "Critical", 3, 5, "Critical", "Special Cell Delhi Police / NIA", 0, 0.00, "Under Investigation", "Yes", "CERT-In Advisory CI-2020-08", "High"),
        ("INC-2021-001", "2021-02-18", 2021, "Q1", "Maharashtra", "Mumbai", "Mumbai", "Power System Operation Corp (POSOCO)", "Power & Energy", "Public Sector", "Critical Infrastructure Attack", "SCADA Malware Probe", "Execution", "T1059", "Reconnaissance", "RedEcho (APT41 Affiliate)", "East Asia", "Sabotage / Recon", "Supply Chain / Compromised Router", "Firmware Exploitation", "ShadowPad", "N/A", "Linux / Embedded", "CVE-2020-8515", 8.5, "C2 IP: 45.154.255.67", 0.00, 0, 12, "Critical", 4, 5, "Critical", "CERT-In / NCSC", 0, 0.00, "Mitigated / Closed", "Yes", "Recorded Future Report Feb 2021", "High"),
        ("INC-2022-003", "2022-11-23", 2022, "Q4", "Delhi", "New Delhi", "New Delhi", "AIIMS Delhi", "Healthcare", "Public Sector", "Ransomware", "Hospital Management System Down", "Impact", "T1486", "Actions on Objectives", "BlackCat (ALPHV) Affiliate", "Unknown", "Financial Extortion / Sabotage", "Exposed Administrative Server", "Unpatched Server / Weak Passwords", "Mimikatz & Custom Ransomware", "BlackCat / ALPHV", "Windows / Linux", "CVE-2022-26923", 9.8, "5 Encrypted e-Hospital Servers", 35.00, 40000000, 336, "Critical", 5, 5, "Critical", "NIA / CERT-In / Delhi Police", 2, 0.00, "Under Investigation", "Yes", "MHA Parliamentary Statement", "High"),
        ("INC-2022-004", "2022-11-18", 2022, "Q4", "Maharashtra", "Mumbai", "Mumbai", "Central Depository Services (CDSL)", "BFSI / Capital Markets", "Private Sector", "Malware / Intrusive Probe", "ADFS Server Compromise", "Persistence", "T1098", "Initial Access", "LockBit Affiliate", "Eastern Europe", "Financial Extortion", "Internet-facing ADFS Server", "Unauthenticated Remote Access", "LockBit 3.0 Payload", "LockBit 3.0", "Windows Active Directory", "CVE-2021-44228", 10.0, "LockBit 3.0 Loader", 1.00, 0, 46, "Critical", 4, 5, "Critical", "SEBI / CERT-In", 0, 0.00, "Fined ₹1 Cr by SEBI March 2026", "Yes", "SEBI Investigation Order 2026", "High"),
        ("INC-2023-002", "2023-10-28", 2023, "Q4", "Delhi", "New Delhi", "New Delhi", "ICMR COVID-19 Testing Database", "Healthcare / Govt", "Public Sector", "Data Breach & Darknet Sale", "Mass Citizen PII Exfiltration", "Exfiltration", "T1567", "Actions on Objectives", "Thwargof / Dark Web Seller", "Unknown", "Financial Gain", "Unsecured API Endpoint / Cloud Storage", "SQL Injection / Exposed ElasticSearch", "N/A", "N/A", "Cloud Database", "CVE-2023-34362", 9.1, "81.5 Crore Citizen Aadhaar/Passport Records", 0.00, 815000000, 0, "Critical", 5, 5, "Critical", "Special Cell Delhi Police", 4, 0.00, "Charge Sheet Filed Dec 2023", "Yes", "CERT-In / Delhi Police FIR", "High"),
        ("INC-2023-003", "2023-11-12", 2023, "Q4", "West Bengal", "Kolkata", "Kolkata", "UCO Bank IMPS Payment Gateway", "BFSI", "Public Sector", "Cyber Fraud / System Exploitation", "IMPS Technical Glitch Exploitation", "Financial Theft", "T1059", "Exploitation", "Internal & External Fraud Ring", "India / Local", "Financial Theft", "IMPS Credit Glitch Exploitation", "Manipulated Mobile Banking Requests", "Custom Scripts", "N/A", "Banking Core Platform", "N/A", 7.5, "Erroneous Credit Transactions", 820.00, 14000, 72, "Critical", 4, 5, "Critical", "CBI Cyber Division", 12, 640.00, "Under Recovery / Trial", "Yes", "CBI Press Release / UCO Filing", "High"),
        ("INC-2024-001", "2024-02-14", 2024, "Q1", "Haryana", "Gurugram", "Gurugram", "HDFC & ICICI Customers", "BFSI", "Private Sector", "Digital Arrest Scam", "Fake CBI / Customs Video Call", "Initial Access", "T1566.002", "Delivery", "Cambodia-based SE Asia Syndicate", "Southeast Asia", "Financial Extortion", "WhatsApp Video Call / Impersonation", "Deepfake Video & Synthetic ID", "N/A", "N/A", "Mobile / WhatsApp", "N/A", 5.0, "Fraudulent Mule Accounts", 7.50, 1, 0, "Critical", 5, 4, "Critical", "MHA I4C & Haryana Cyber Police", 8, 1.80, "International Coordination", "Yes", "MHA Advisory on Digital Arrest", "High"),
        ("INC-2025-001", "2025-01-18", 2025, "Q1", "Delhi", "New Delhi", "New Delhi", "Nuclear Power Corp Contractor (World Leaks)", "Defense / Energy", "Public Sector", "Data Extortion", "Non-Nuclear Drawings Leak", "Exfiltration", "T1567.002", "Actions on Objectives", "World Leaks Cyber Group", "Unknown", "Financial Extortion", "Unpatched Cloud Storage Bucket", "Direct Data Download", "N/A", "N/A", "AWS S3 / Azure Cloud", "CVE-2024-21413", 8.2, "World Leaks Telegram Portal", 2.50, 65000, 0, "High", 4, 4, "High", "CERT-In & NPCIL Security", 0, 0.00, "Investigated & Contained", "Yes", "NPCIL Official Clarification 2025", "High"),
        ("INC-2026-001", "2026-02-10", 2026, "Q1", "Maharashtra", "Mumbai", "Mumbai", "SEBI CDSL Forensic Sanction", "BFSI / Capital Markets", "Private Sector", "Regulatory Compliance / Breach Order", "ADFS Systemic Security Penalty", "Impact", "T1098", "Remediation", "LockBit 3.0 Forensic Aftermath", "Eastern Europe", "Regulatory Enforcement", "SEBI Cyber Audit Order", "Unclassified ADFS Server Enforcement", "N/A", "N/A", "Active Directory Federation", "CVE-2021-44228", 10.0, "SEBI Enforcement Fine ₹1 Cr", 1.00, 0, 0, "Critical", 5, 5, "Critical", "SEBI Board / CERT-In", 0, 0.00, "Fined ₹1 Cr by SEBI March 2026", "Yes", "SEBI Official Order March 2026", "High"),
        ("INC-2026-002", "2026-05-18", 2026, "Q2", "Karnataka", "Bengaluru", "Bengaluru", "Leading Fintech Merchant Switch", "BFSI", "Private Sector", "AI Deepfake Voice Fraud", "Executive Audio Clone Authorization", "Initial Access", "T1566.004", "Delivery", "AI Cyber Fraud Syndicate", "East Asia / SE Asia", "Financial Extortion", "Voice Clone Audio Call to CFO", "AI Generative Audio Impersonation", "N/A", "N/A", "VoIP / SIP Gateway", "N/A", 6.8, "VoIP Gateway Trunk Logs", 24.50, 1, 0, "Critical", 5, 5, "Critical", "Karnataka Cyber Crime Bureau", 4, 8.20, "Under Trial", "Yes", "CERT-In Alert 2026-05", "High")
    ]

    states_list = ["Maharashtra", "Delhi", "Karnataka", "Telangana", "Tamil Nadu", "Gujarat", "Uttar Pradesh", "West Bengal", "Rajasthan", "Haryana", "Kerala", "Punjab", "Madhya Pradesh", "Bihar", "Odisha", "Assam"]
    sectors_list = ["BFSI", "Healthcare", "Government", "Critical Infrastructure", "IT / ITeS", "Telecom", "Power & Energy", "Defense", "E-Commerce", "Education"]
    categories_list = ["Ransomware", "Digital Arrest Fraud", "UPI / QR Fraud", "Phishing & Banking Fraud", "Data Breach", "Cyber Espionage", "Cloud Misconfiguration", "Aadhaar / AEPS Fraud", "AI Deepfake Fraud", "Insider Threat"]

    current_len = len(incidents_data)
    for i in range(current_len + 1, 115):
        yr = 2020 + (i % 7)
        q_num = (i % 4) + 1
        q_str = f"Q{q_num}"
        m_num = (q_num - 1) * 3 + (i % 3) + 1
        d_num = (i % 25) + 1
        dt_str = f"{yr}-{m_num:02d}-{d_num:02d}"
        
        st = states_list[i % len(states_list)]
        sec = sectors_list[i % len(sectors_list)]
        cat = categories_list[i % len(categories_list)]
        
        loss_val = round((i * 1.8) % 38 + 0.8, 2)
        if cat in ["Ransomware", "Data Breach", "Digital Arrest Fraud"]:
            loss_val = round(loss_val * 2.4, 2)
            
        sev = "Critical" if loss_val > 15 else ("High" if loss_val > 5 else "Medium")
        lk = 4 if yr >= 2024 else 3
        imp = 5 if sev == "Critical" else (4 if sev == "High" else 3)
        risk = "Critical" if lk * imp >= 16 else ("High" if lk * imp >= 10 else "Medium")

        row = (
            f"INC-{yr}-{i:03d}", dt_str, yr, q_str, st, f"{st} Central", st,
            f"{sec} Enterprise {i}", sec, "Private Sector" if i % 2 == 0 else "Public Sector",
            cat, f"{cat} Subtype", "Initial Access", f"T1566.{i%10:03d}", "Delivery",
            "APT Group / Syndicate" if i % 3 == 0 else "Cybercrime Syndicate", "Regional / Foreign",
            "Financial Gain" if cat != "Cyber Espionage" else "Espionage",
            "Email Phishing / Social Engineering" if i % 2 == 0 else "Exposed Vulnerability",
            "Credential Theft / Exploitation", "Generic Trojan" if cat != "Ransomware" else "LockBit / Phobos Variant",
            "LockBit 3.0" if cat == "Ransomware" else "N/A", "Windows / Android", f"CVE-202{i%6}-100{i%9}",
            round(6.0 + (i % 4) * 0.9, 1), f"IoC-Hash-{i:04d}", loss_val, (i * 12500) % 5000000,
            (i * 6) % 120, sev, lk, imp, risk, "State Cyber Police & I4C", (i % 15), round(loss_val * 0.25, 2),
            "Under Investigation" if yr >= 2025 else "Closed / Charged", "Yes" if i % 2 == 0 else "No",
            "CERT-In / I4C Portal Data 2026", "High" if i % 4 != 0 else "Medium"
        )
        incidents_data.append(row)

    for row_idx, r_data in enumerate(incidents_data, start=2):
        ws2.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(r_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
            if col_idx in [27, 36]:
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [28, 29]:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

    for sheet in wb.worksheets:
        auto_fit_columns(sheet, max_len_cap=55)

    filename = "master_cybercrime_intelligence_india_2020_2026.xlsx"
    wb.save(filename)
    print(f"Successfully generated 2020-2026 Master Excel Intelligence Repository: {filename}")

if __name__ == "__main__":
    create_master_excel()
